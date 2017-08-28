# -*- coding: UTF-8 -*-
import openpyxl,csv,os

exl=openpyxl.Workbook()

for csvfile in os.listdir("."):
    if not csvfile.endswith(".csv"):
        continue
    print('change csv file ' + csvfile + '...')
    csvrow=[]
    file= open(csvfile)
    pageobj= csv.reader(file)
    for i in pageobj:
        csvrow.append(i)
    file.close()
    print(csvrow)

    exlname=csvfile.split('.')[0]
    exl.create_sheet(exlname)
    sheet=exl.get_sheet_by_name(exlname)
    for num in range(1,len(csvrow)+1):
        for inum in range(1,len(csvrow[num-1])+1):
            sheet.cell(row=num,column=inum).value=csvrow[num-1][inum-1]

exl.save("update.xlsx")


