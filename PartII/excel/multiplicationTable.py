# -*- coding: UTF-8 -*-
import openpyxl

exl = openpyxl.load_workbook("result.xlsx")
sheet = exl.get_sheet_by_name("Sheet1")
a= input("enter a num:")
for i in range(1,int(a)+1):
    sheet.cell(row=1,column=i+1).value=i
    sheet.cell(row=i+1,column= 1).value=i
    for j in range(2,int(a)+1):
        sheet.cell()

exl.save("result.xlsx")