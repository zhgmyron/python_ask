# -*- coding: UTF-8 -*-
import openpyxl
exl= openpyxl.load_workbook("result.xlsx")
sheet= exl.get_sheet_by_name("Sheet2")
r= sheet.max_row
c= sheet.max_column
temp_list=[]
for a in range(1,r+1):
    for b in range(1,c+1):
        temp_list.append(sheet.cell(row=a,column=b).value)
        sheet.cell(row=a,column=b).value=''
temp=temp_list[::-1]
for a in range(1,r+1):
    for b in range(1,c+1):
        sheet.cell(row=b,column=a).value=temp.pop()
exl.save('r.xlsx')