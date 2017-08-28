# -*- coding: UTF-8 -*-
import openpyxl
def open_excel(bookname):
    exl= openpyxl.load_workbook(bookname,data_only=True)
    sheet= exl.get_sheet_by_name('Sheet')
    for b in range(1,sheet.max_column+1):
        w=open_file(str(b)+'.txt')
        for a in range(1,sheet.max_row+1):
            temp= sheet.cell(row=a,column=b).value
            w.write(str(temp)+',')
def open_file(filename):
    wb= open(filename,'w')
    return wb
open_excel('produceSales.xlsx')