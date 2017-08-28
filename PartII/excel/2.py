# -*- coding: UTF-8 -*-
import openpyxl
start_line= int(input("enter start:"))
m=int(input("blank:"))
exl= openpyxl.load_workbook("produceSales.xlsx",data_only=True)
sheet = exl.get_sheet_by_name("Sheet")
r=sheet.max_row
c= sheet.max_column
temp_list=[]
for a in range(start_line,r+1):
    for b in range(1,c+1):
        temp_list.append(sheet.cell(row=a,column=b).value)
        sheet.cell(row=a,column=b).value=''
print(temp_list[:20])
temp= temp_list[::-1]

for a in range(start_line+m,r+1+m):
    for b in range(1,c+1):
        sheet.cell(row=a,column=b).value= temp.pop()

exl.save("new.xlsx")
