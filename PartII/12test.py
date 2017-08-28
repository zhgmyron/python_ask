# -*- coding: UTF-8 -*-
import openpyxl

wb =openpyxl.load_workbook('jobstatus.xlsx')

#print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')
print (sheet['C1'].value)
print (sheet.cell(row=1,column=2).coordinate)

a = sheet.max_row
b= sheet.max_column

print (a)
print(sheet['A1':'C3'])
#for cell in s